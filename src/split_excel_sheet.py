from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
import argparse
import os

def copyRange(startCol, startRow, endCol, endRow, sheet):
    '''
    Copies a range of cells from an excel workbook's sheet

    Parameters
    ----------
    startCol : int
        col to start copying inside `sheet`
    startRow : int
        row to start copying inside `sheet`
    endCol : int
        col to stop copying inside `sheet`
    endRow : int
        row to stop copying inside `sheet`
    sheet : derived Workbook object
        pointer to active sheet in working excel file
    
    Returns
    -------
    rangeSelected : array_like
        list of objects extracted from `sheet` cells

    '''
    rangeSelected = []
    #Loops through selected Rows
    for i in range(startRow,endRow + 1):
        #Appends the row to a RowSelected list
        rowSelected = []
        for j in range(startCol,endCol+1):
            rowSelected.append(sheet.cell(row = i, column = j).value)
        #Adds the RowSelected List and nests inside the rangeSelected
        rangeSelected.append(rowSelected)
 
    return rangeSelected

def pasteRange(startCol, startRow, endCol, endRow, sheetReceiving, copiedData):
    '''
    Copies a range of cells from an excel workbook's sheet

    Parameters
    ----------
    startCol : int
        col to start copying inside `sheet`
    startRow : int
        row to start copying inside `sheet`
    endCol : int
        col to stop copying inside `sheet`
    endRow : int
        row to stop copying inside `sheet`
    sheetReceiving : derived Workbook object
        pointer to sheet in excel workbook copying to
    copiedData : array_like
        list of data to be copied to receiving sheet

    '''
    for countRow, i in enumerate(range(startRow,endRow+1)):
        for countCol, j in enumerate(range(startCol,endCol+1)):
            sheetReceiving.cell(row = i, column = j).value = copiedData[countRow][countCol]

def split_workbook(file, dir_out, split_size):
    '''
    Main workhorse that has smarts to split workbook

    Parameters
    ----------
    file : str
        xlsx file to read in and split up
    dir_out : str
        directory to save split files into
    split_size : number of rows to split rows of `file` by

    '''

    # Check `file` exists, if not return
    if not os.path.exists(file):
        return

    # Check if `dir_out` exists, if not create it
    if not os.path.exists(dir_out):
        os.makedirs(dir_out)

    # Load workbook
    wb = load_workbook(file)

    # Strip file name for saving purposes
    out_file = os.path.splitext(os.path.basename(file))[0]

    # Get first sheet
    sheet = wb.worksheets[0]

    # Get num of cols
    max_cols = sheet.max_column-1 # there is a header

    # Get max num of rows
    total_rows = sheet.max_row

    # Get number of cols each subfile should hold, and remainder, -1 because of header
    (num_files, remainder) = divmod(total_rows-1, split_size)
    if remainder > 0:
        print('Warning: the last subfile will have {} + {} rows'.format(split_size, remainder))
    
    # Create subfiles containing only split_size number of cols of original
    for subfile in range(num_files):
        # Create new workbook name
        dest_filename = '{}_{}.xlsx'.format(os.path.join(dir_out,out_file), subfile)
        # Create new workbook
        new_file = Workbook()
        new_sheet = new_file.active
        # Get data to copy, after first everything has +1 row due to header
        if subfile == 0:
            start_row = 2
            end_row = subfile*split_size+split_size+1
        elif subfile == num_files-1:
            start_row = subfile*split_size+2
            end_row = subfile*split_size+split_size+remainder+1
        else:
            start_row = subfile*split_size+2
            end_row = subfile*split_size+split_size+1
        # Copy header
        header = copyRange(1, 1, max_cols, 1, sheet)
        # Paste header
        pasteRange(1, 1, max_cols, 1, new_sheet, header)
        # Copy data
        data =  copyRange(1, start_row, max_cols, end_row, sheet)
        # Paste data
        if subfile == num_files-1:
            pasteRange(1, 2, max_cols, split_size+remainder+1, new_sheet, data)
        else:
            pasteRange(1, 2, max_cols, split_size+1, new_sheet, data)
        print('Filled subfile {}'.format(subfile))
        new_file.save(filename = dest_filename)


if __name__ == '__main__':
    # Set up argument parser
    parser = argparse.ArgumentParser(description='Splits 1 excel file into X smaller files, as equally as possible based on # of columns')
    parser.add_argument('file', help='Excel file to split')
    parser.add_argument('dir_out', help='Directory to save split files to')
    parser.add_argument('split_size', help='Number of rows to split by')

    # Get arguments
    args = parser.parse_args()

    # Save arguments locally
    file = args.file
    dir_out = args.dir_out
    split_size = args.split_size

    split_workbook(file, dir_out, int(split_size))

