from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
import argparse
import os

#Copy range of cells as a nested list
#Takes: start cell, end cell, and sheet you want to copy from.
#src: https://yagisanatode.com/2017/11/18/copy-and-paste-ranges-in-excel-with-openpyxl-and-python-3/
def copyRange(startCol, startRow, endCol, endRow, sheet):
    rangeSelected = []
    #Loops through selected Rows
    for i in range(startRow,endRow + 1):
        #Appends the row to a RowSelected list
        rowSelected = []
        for j in range(startCol,endCol+1):
            rowSelected.append(sheet.cell(row = i, column = j).value)
        #Adds the RowSelected List and nests inside the rangeSelected
        rangeSelected.append(rowSelected)
 
    return rangeSelected

#Paste range
#Paste data from copyRange into template sheet
#src: https://yagisanatode.com/2017/11/18/copy-and-paste-ranges-in-excel-with-openpyxl-and-python-3/
def pasteRange(startCol, startRow, endCol, endRow, sheetReceiving,copiedData):
    countRow = 0
    print(copiedData)
    for i in range(startRow,endRow+1):
        countCol = 0
        for j in range(startCol,endCol+1):
            print(countRow, countCol)
            sheetReceiving.cell(row = i, column = j).value = copiedData[countRow][countCol]
            countCol += 1
        countRow += 1

def split_workbook(file, dir_out, split_size):
    # Load workbook
    wb = load_workbook(file)

    # Strip file name for saving purposes
    out_file = os.path.splitext(os.path.basename(file))[0]

    # Get first sheet
    sheet = wb.worksheets[0]

    # Get num of cols
    max_cols = sheet.max_column-1 # there is a header

    # Get max num of rows
    total_rows = sheet.max_row

    # Get number of cols each subfile should hold, and remainder, -1 because of header
    (num_files, remainder) = divmod(total_rows-1, split_size)
    print(num_files, remainder)
    if remainder > 0:
        print('Warning: the last subfile will have {} + {} rows'.format(split_size, remainder))
    
    # Create subfiles containing only split_size number of cols of original
    for subfile in range(num_files):
        # Create new workbook name
        dest_filename = '{}_{}.xlsx'.format(os.path.join(dir_out,out_file), subfile)
        # Create new workbook
        new_file = Workbook()
        new_sheet = new_file.active
        # Get data to copy, after first everything has +1 row due to header
        if subfile == 0:
            start_row = 2
            end_row = subfile*split_size+split_size+1
        elif subfile == num_files-1:
            start_row = subfile*split_size+2
            end_row = subfile*split_size+split_size+remainder+1
        else:
            start_row = subfile*split_size+2
            end_row = subfile*split_size+split_size+1
        # Copy header
        header = copyRange(1, 1, max_cols, 1, sheet)
        # Paste header
        pasteRange(1, 1, max_cols, 1, new_sheet, header)
        # Copy data
        data =  copyRange(1, start_row, max_cols, end_row, sheet)
        # Paste data
        if subfile == num_files-1:
            pasteRange(1, 2, max_cols, split_size+remainder+1, new_sheet, data)
        else:
            pasteRange(1, 2, max_cols, split_size+1, new_sheet, data)
        print('Filled subfile {}'.format(subfile))
        new_file.save(filename = dest_filename)


if __name__ == '__main__':
    # Set up argument parser
    parser = argparse.ArgumentParser(description='Splits 1 excel file into X smaller files, as equally as possible based on # of columns')
    parser.add_argument('file', help='Excel file to split')
    parser.add_argument('dir_out', help='Directory to save split files to')
    parser.add_argument('split_size', help='Number of rows to split by')

    # Get arguments
    args = parser.parse_args()

    # Save arguments locally
    file = args.file
    dir_out = args.dir_out
    split_size = args.split_size

    split_workbook(file, dir_out, int(split_size))

